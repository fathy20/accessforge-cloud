from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .schemas import CrewHoursReportResponse, CrewMemberSummary, FlightItem


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DETAIL_HEADERS = (
    "Position type",
    "Name",
    "Date",
    "Aircraft",
    "Flight number",
    "ADEP",
    "ADES",
    "OFF",
    "ON",
    "Block time",
    "Augmented (Heavy)",
    "Heavy Source",
    "Heavy Reason",
    "LEON Heavy",
    "Derived Heavy",
    "Training (TRN)",
    "Unknown Resolution",
)

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
_DURATION_PATTERN = re.compile(
    r"^\s*(?P<hours>\d+):(?P<minutes>[0-5]\d)(?::(?P<seconds>[0-5]\d))?\s*$"
)

_TITLE_FILL = PatternFill(fill_type="solid", fgColor="17365D")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
_SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
_LABEL_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")
_SEPARATOR_FILL = PatternFill(fill_type="solid", fgColor="F3F6F9")
_WHITE_FONT = Font(color="FFFFFF", bold=True, size=14)
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SECTION_FONT = Font(color="17365D", bold=True)
_LABEL_FONT = Font(bold=True, color="17365D")
_THIN_SIDE = Side(style="thin", color="B7C9D6")
_MEDIUM_SIDE = Side(style="medium", color="5B7890")
_CELL_BORDER = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=_THIN_SIDE,
    bottom=_THIN_SIDE,
)
_SEPARATOR_BORDER = Border(top=_MEDIUM_SIDE, bottom=_MEDIUM_SIDE)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)


def _write_string(cell: Cell, value: str | None) -> None:
    """Write text safely so spreadsheet formulas cannot be injected."""

    if value is None:
        cell.value = None
        return
    if value.startswith(_FORMULA_PREFIXES):
        value = f"'{value}"
    cell.value = value


def _summary_position(position_type: str | None) -> str:
    return position_type or "Unclassified"


def _aircraft_display(flight: FlightItem) -> str:
    registration = flight.aircraft_reg or ""
    aircraft_type = flight.aircraft_type or ""
    if registration and aircraft_type:
        return f"{registration} ({aircraft_type})"
    return registration or aircraft_type


def _parse_duration(value: str) -> float | None:
    match = _DURATION_PATTERN.fullmatch(value)
    if match is None:
        return None
    hours = int(match.group("hours"))
    minutes = int(match.group("minutes"))
    seconds = int(match.group("seconds") or 0)
    return (hours * 60 * 60 + minutes * 60 + seconds) / (24 * 60 * 60)


def _write_duration(cell: Cell, value: str | None, unparsed: list[str]) -> None:
    if value is None:
        _write_string(cell, None)
        return
    parsed = _parse_duration(value)
    if parsed is None:
        _write_string(cell, value)
        if value:
            unparsed.append(value)
        return
    cell.value = parsed
    cell.number_format = "[h]:mm"


def _augmented_heavy_display(value: bool | None) -> str:
    if value is True:
        return "Yes"
    if value is False:
        return "No"
    return "Unknown"


def _training_source_display(flight: FlightItem) -> str | None:
    """Say where the trainee flag came from: the role slot or the Work Schedule."""

    sources = []
    if flight.is_training_position:
        sources.append("Position")
    if flight.is_training_function:
        sources.append("Function")
    if not sources and flight.is_trn:
        sources.append("MCP total")
    return " + ".join(sources) or None


def _style_cell(cell: Cell, *, alignment: Alignment = _CELL_ALIGNMENT) -> None:
    cell.border = _CELL_BORDER
    cell.alignment = alignment


def _style_merged_row(
    worksheet,
    row: int,
    column_count: int,
    *,
    fill: PatternFill,
    font: Font,
    alignment: Alignment,
) -> None:
    for column in range(1, column_count + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = _CELL_BORDER


def _write_table_headers(worksheet, row: int, headers: tuple[str, ...]) -> None:
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=column)
        _write_string(cell, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _CELL_BORDER
        cell.alignment = _HEADER_ALIGNMENT
    worksheet.row_dimensions[row].height = 30


def _write_label_value(worksheet, row: int, label: str, value: str | None) -> None:
    label_cell = worksheet.cell(row=row, column=1)
    value_cell = worksheet.cell(row=row, column=2)
    _write_string(label_cell, label)
    _write_string(value_cell, value)
    label_cell.fill = _LABEL_FILL
    label_cell.font = _LABEL_FONT
    label_cell.alignment = _CELL_ALIGNMENT
    value_cell.alignment = _CELL_ALIGNMENT
    label_cell.border = _CELL_BORDER
    value_cell.border = _CELL_BORDER


def _write_official_total(cell: Cell, value: str | None, unparsed: list[str]) -> None:
    if value is None or value == "":
        _write_string(cell, "Not available")
        return
    if value == "TRN":
        _write_string(cell, value)
        return
    _write_duration(cell, value, unparsed)


def _write_duration_label_value(
    worksheet,
    row: int,
    label: str,
    value: str | None,
    unparsed: list[str],
) -> None:
    _write_label_value(worksheet, row, label, None)
    _write_official_total(worksheet.cell(row=row, column=2), value, unparsed)


def _configure_detail_sheet(worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"
    worksheet.print_title_rows = "4:4"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)
    widths = (16, 24, 14, 18, 16, 12, 12, 22, 22, 14, 20, 20, 24, 14, 16, 18, 32)
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _configure_summary_sheet(worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A4"
    worksheet.print_title_rows = "3:3"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(left=0.35, right=0.35, top=0.5, bottom=0.5)
    widths = (30, 20, 14, 18)
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _configure_information_sheet(worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 85
    worksheet.page_margins = PageMargins(left=0.35, right=0.35, top=0.5, bottom=0.5)


def _add_detail_crew_block(
    worksheet,
    row: int,
    crew: CrewMemberSummary,
    report: CrewHoursReportResponse,
    unparsed_durations: list[str],
) -> int:
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DETAIL_HEADERS))
    _write_string(worksheet.cell(row=row, column=1), f"Crew member: {crew.display_name}")
    _style_merged_row(
        worksheet,
        row,
        len(DETAIL_HEADERS),
        fill=_SECTION_FILL,
        font=_SECTION_FONT,
        alignment=_CELL_ALIGNMENT,
    )
    worksheet.row_dimensions[row].height = 22
    row += 1
    _write_label_value(
        worksheet,
        row,
        "Report period",
        f"{report.period.from_date} to {report.period.to_date}",
    )
    row += 1
    _write_duration_label_value(
        worksheet,
        row,
        "Official total",
        crew.official_total,
        unparsed_durations,
    )
    row += 1

    for flight in crew.flights:
        values = (
            flight.position,
            crew.display_name,
            flight.flight_date,
            _aircraft_display(flight),
            flight.flight_number,
            flight.departure_airport,
            flight.arrival_airport,
            flight.start_time_utc,
            flight.end_time_utc,
        )
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row, column=column)
            _write_string(cell, value)
            _style_cell(cell)
        duration_cell = worksheet.cell(
            row=row,
            column=DETAIL_HEADERS.index("Block time") + 1,
        )
        _write_duration(duration_cell, flight.block_time, unparsed_durations)
        _style_cell(duration_cell)
        augmented_column = DETAIL_HEADERS.index("Augmented (Heavy)") + 1
        augmented_cell = worksheet.cell(row=row, column=augmented_column)
        _write_string(augmented_cell, _augmented_heavy_display(flight.augmented_heavy))
        _style_cell(augmented_cell)
        provenance_values = (
            flight.heavy_source,
            flight.heavy_reason,
            _augmented_heavy_display(flight.leon_heavy),
            _augmented_heavy_display(flight.derived_heavy),
            _training_source_display(flight),
            # Keyed on the reason, not on the badge: the badge now means only
            # "the resolver established Heavy", while the reason is present for
            # every leg STEP 4 touched. This keeps the exported column exactly
            # as it was before the badge rule narrowed.
            flight.unknown_resolution_reason,
        )
        for column, value in enumerate(provenance_values, start=augmented_column + 1):
            provenance_cell = worksheet.cell(row=row, column=column)
            _write_string(provenance_cell, value)
            _style_cell(provenance_cell)
        worksheet.row_dimensions[row].height = 22
        row += 1
    return row


def _add_separator(worksheet, row: int, column_count: int) -> None:
    for column in range(1, column_count + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.fill = _SEPARATOR_FILL
        cell.border = _SEPARATOR_BORDER
    worksheet.row_dimensions[row].height = 8


def _build_detail_sheet(
    worksheet,
    report: CrewHoursReportResponse,
    position_type: str,
    unparsed_durations: list[str],
) -> None:
    _configure_detail_sheet(worksheet)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DETAIL_HEADERS))
    _write_string(worksheet.cell(row=1, column=1), f"Crew Hours — {position_type} Detail")
    _style_merged_row(
        worksheet,
        1,
        len(DETAIL_HEADERS),
        fill=_TITLE_FILL,
        font=_WHITE_FONT,
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    worksheet.row_dimensions[1].height = 26
    _write_label_value(
        worksheet,
        2,
        "Report period",
        f"{report.period.from_date} to {report.period.to_date}",
    )
    _write_label_value(worksheet, 3, "Source", "LEON MCP")
    _write_table_headers(worksheet, 4, DETAIL_HEADERS)

    crews = [crew for crew in report.crew_members if crew.position_type == position_type]
    row = 5
    for index, crew in enumerate(crews):
        if index:
            _add_separator(worksheet, row, len(DETAIL_HEADERS))
            row += 1
        row = _add_detail_crew_block(worksheet, row, crew, report, unparsed_durations)

    last_row = max(row - 1, 4)
    worksheet.auto_filter.ref = f"A4:{get_column_letter(len(DETAIL_HEADERS))}{last_row}"


def _build_summary_sheet(
    worksheet,
    report: CrewHoursReportResponse,
    unparsed_durations: list[str],
) -> None:
    _configure_summary_sheet(worksheet)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    _write_string(worksheet.cell(row=1, column=1), "Crew Hours — Summary")
    _style_merged_row(
        worksheet,
        1,
        4,
        fill=_TITLE_FILL,
        font=_WHITE_FONT,
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    worksheet.row_dimensions[1].height = 26
    _write_label_value(
        worksheet,
        2,
        "Report period",
        f"{report.period.from_date} to {report.period.to_date}",
    )
    _write_table_headers(worksheet, 3, ("Name", "Position type", "Flight count", "Official total"))

    for row, crew in enumerate(report.crew_members, start=4):
        values = (
            crew.display_name,
            _summary_position(crew.position_type),
        )
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row, column=column)
            _write_string(cell, value)
            _style_cell(cell)
        flight_count_cell = worksheet.cell(row=row, column=3)
        flight_count_cell.value = crew.flight_count
        _style_cell(flight_count_cell, alignment=Alignment(horizontal="center", vertical="center"))
        official_total_cell = worksheet.cell(row=row, column=4)
        _write_official_total(official_total_cell, crew.official_total, unparsed_durations)
        _style_cell(official_total_cell)

    worksheet.auto_filter.ref = f"A3:D{max(len(report.crew_members) + 3, 3)}"


def _utc_iso(generated_at: datetime) -> str:
    if generated_at.tzinfo is None:
        generated_at = generated_at.replace(tzinfo=timezone.utc)
    return generated_at.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def _build_information_sheet(
    worksheet,
    report: CrewHoursReportResponse,
    generated_at: datetime,
    generated_by: str,
    unparsed_durations: list[str],
) -> None:
    _configure_information_sheet(worksheet)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    _write_string(worksheet.cell(row=1, column=1), "Crew Hours — Report Information")
    _style_merged_row(
        worksheet,
        1,
        2,
        fill=_TITLE_FILL,
        font=_WHITE_FONT,
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    worksheet.row_dimensions[1].height = 26
    _write_label_value(worksheet, 3, "Report period from", report.period.from_date)
    _write_label_value(worksheet, 4, "Report period to", report.period.to_date)
    _write_label_value(worksheet, 5, "Generated at UTC", _utc_iso(generated_at))
    _write_label_value(worksheet, 6, "Generated by", generated_by)
    _write_label_value(worksheet, 7, "Source", "LEON MCP")
    _write_label_value(
        worksheet,
        8,
        "Crew handling",
        "Maintenance and Unclassified crew members are included on Summary; detail sheets contain Cockpit and Cabin only.",
    )
    unparsed = list(dict.fromkeys(unparsed_durations))
    duration_note = (
        "None"
        if not unparsed
        else "Unparsed duration strings were preserved as text: " + ", ".join(unparsed)
    )
    _write_label_value(worksheet, 9, "Unparsed durations", duration_note)


def build_crew_hours_workbook(
    report: CrewHoursReportResponse,
    *,
    generated_at: datetime | None = None,
    generated_by: str = "",
) -> BytesIO:
    generated_at = generated_at or datetime.now(timezone.utc)
    workbook = Workbook()
    cockpit_sheet = workbook.active
    cockpit_sheet.title = "Cockpit Detail"
    cabin_sheet = workbook.create_sheet("Cabin Detail")
    summary_sheet = workbook.create_sheet("Summary")
    information_sheet = workbook.create_sheet("Report Information")

    unparsed_durations: list[str] = []
    _build_detail_sheet(cockpit_sheet, report, "Cockpit", unparsed_durations)
    _build_detail_sheet(cabin_sheet, report, "Cabin", unparsed_durations)
    _build_summary_sheet(summary_sheet, report, unparsed_durations)
    _build_information_sheet(
        information_sheet,
        report,
        generated_at,
        generated_by,
        unparsed_durations,
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _safe_filename_component(value: str) -> str:
    collapsed = re.sub(r"[^A-Za-z0-9._-]+", "-", value)
    return collapsed.strip("-") or "unknown"


def build_crew_hours_filename(report: CrewHoursReportResponse, _generated_at: datetime) -> str:
    from_date = _safe_filename_component(report.period.from_date)
    to_date = _safe_filename_component(report.period.to_date)
    return f"crew-hours-{from_date}-to-{to_date}.xlsx"
