import os
import re
import shutil
import sys
import tempfile
import unittest
from datetime import timedelta
from pathlib import Path
from types import SimpleNamespace

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from fastapi.testclient import TestClient
from openpyxl import load_workbook

_API_TEMP_DIR = Path(tempfile.mkdtemp(prefix="crew_hours_export_api_"))
_API_TEMP_DB = _API_TEMP_DIR / "api.db"
_original_database_url = os.environ.get("DATABASE_URL")
_original_jwt_secret = os.environ.get("JWT_SECRET_KEY")
os.environ["DATABASE_URL"] = f"sqlite:///{_API_TEMP_DB.as_posix()}"
os.environ["JWT_SECRET_KEY"] = _original_jwt_secret or "test-secret-key-for-ci-only-32-chars"
# Keep this API test deterministic: never load live LEON credentials during collection.
for _name in ("LEON_BASE_URL", "LEON_REFRESH_TOKEN", "LEON_MCP_URL", "LEON_TIMEOUT_SECONDS"):
    os.environ[_name] = ""

from backend.auth import get_current_user
from backend.main import CORS_ORIGINS, app
from backend.statistics.crew_hours.errors import LeonRateLimitError, LeonTimeoutError
from backend.statistics.crew_hours.schemas import (
    CrewHoursPeriod,
    CrewHoursReportResponse,
    CrewMemberSummary,
    FlightItem,
)
from backend.statistics.crew_hours.service import get_crew_hours_service

if _original_database_url is None:
    os.environ.pop("DATABASE_URL", None)
else:
    os.environ["DATABASE_URL"] = _original_database_url
if _original_jwt_secret is None:
    os.environ.pop("JWT_SECRET_KEY", None)
else:
    os.environ["JWT_SECRET_KEY"] = _original_jwt_secret


def _flight(
    flight_nid: str,
    *,
    position: str,
    block_time: str,
    flight_number: str,
    journey_log: dict | None = None,
) -> FlightItem:
    return FlightItem(
        flight_nid=flight_nid,
        flight_number=flight_number,
        departure_airport="CAI",
        arrival_airport="HRG",
        start_time_utc="2026-06-03T06:00:00Z",
        end_time_utc="2026-06-03T07:15:00Z",
        aircraft_reg="SU-RED",
        aircraft_type="A320",
        flight_date="2026-06-03",
        block_time=block_time,
        position=position,
        journey_log=journey_log,
    )


def _report() -> CrewHoursReportResponse:
    return CrewHoursReportResponse(
        period=CrewHoursPeriod(from_date="2026-06-01", to_date="2026-06-30"),
        source="leon_mcp_report",
        hours_source_status="official_mcp_report",
        total_crew=4,
        total_flights=3,
        records_count=3,
        official_totals_available=2,
        official_totals_unavailable=2,
        official_totals_by_position={"Cockpit": "75:05", "Cabin": "01:30"},
        crew_members=[
            CrewMemberSummary(
                crew_id="INTERNAL-CREW-001",
                person_code="INTERNAL-CODE-001",
                display_name="=cmd|' /c calc'!A0",
                full_name="Formula test name",
                position_type="Cockpit",
                official_total="75:05",
                flight_count=2,
                flights=[
                    _flight(
                        "INTERNAL-FLIGHT-001",
                        position="PAD",
                        block_time="01:30",
                        flight_number="RS101",
                        journey_log={"raw": "RAW LEON PAYLOAD"},
                    ),
                    _flight(
                        "INTERNAL-FLIGHT-002",
                        position="CPT",
                        block_time="01:30",
                        flight_number="RS102",
                    ),
                ],
            ),
            CrewMemberSummary(
                crew_id="INTERNAL-CREW-002",
                person_code="INTERNAL-CODE-002",
                display_name="Cabin Crew",
                full_name="Cabin Crew",
                position_type="Cabin",
                official_total=None,
                flight_count=1,
                flights=[
                    _flight(
                        "INTERNAL-FLIGHT-003",
                        position="FA1",
                        block_time="not-a-duration",
                        flight_number="RS103",
                    ),
                ],
            ),
            CrewMemberSummary(
                crew_id="INTERNAL-CREW-003",
                person_code="INTERNAL-CODE-003",
                display_name="Maintenance Crew",
                position_type="Maintenance",
                official_total="01:00",
                flight_count=0,
                flights=[],
            ),
            CrewMemberSummary(
                crew_id="INTERNAL-CREW-004",
                person_code="INTERNAL-CODE-004",
                display_name="Unclassified Crew",
                position_type=None,
                official_total="",
                flight_count=0,
                flights=[],
            ),
        ],
    )


class _RecordingReportService:
    def __init__(self, result=None, error=None):
        self.result = result
        self.error = error
        self.calls = 0
        self.arguments = []

    def get_crew_hours_report(self, **kwargs):
        self.calls += 1
        self.arguments.append(kwargs)
        if self.error is not None:
            raise self.error
        return self.result


class TestCrewHoursExport(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.client = TestClient(app)

    def setUp(self):
        self.service = _RecordingReportService(result=_report())
        app.dependency_overrides[get_current_user] = lambda: SimpleNamespace(
            email="exporter@example.com"
        )
        app.dependency_overrides[get_crew_hours_service] = lambda: self.service

    def tearDown(self):
        app.dependency_overrides.pop(get_current_user, None)
        app.dependency_overrides.pop(get_crew_hours_service, None)

    @classmethod
    def tearDownClass(cls):
        import backend.database as database

        database.engine.dispose()
        shutil.rmtree(_API_TEMP_DIR, ignore_errors=True)

    def _export_response(self, **params):
        query = {"from": "2026-06-01", "to": "2026-06-30", **params}
        return self.client.get("/api/statistics/crew-hours/report/export", params=query)

    def test_endpoint_requires_authentication(self):
        app.dependency_overrides.pop(get_current_user, None)

        response = self._export_response()

        self.assertEqual(response.status_code, 401)
        self.assertEqual(self.service.calls, 0)

    def test_query_validation_matches_report_endpoint_and_skips_service(self):
        for query, expected_detail in (
            (
                {"from": "2026-13-01", "to": "2026-06-30"},
                "Query parameter 'from' must be a valid YYYY-MM-DD date.",
            ),
            (
                {"from": "2026-06-30", "to": "2026-06-01"},
                "Query parameter 'from' must not be after 'to'.",
            ),
        ):
            with self.subTest(query=query):
                response = self.client.get(
                    "/api/statistics/crew-hours/report/export",
                    params=query,
                )
                self.assertEqual(response.status_code, 422)
                self.assertEqual(response.json(), {"detail": expected_detail})
                self.assertEqual(self.service.calls, 0)

    def test_same_service_method_feeds_screen_and_export_without_live_leon(self):
        report_response = self.client.get(
            "/api/statistics/crew-hours/report",
            params={"from": "2026-06-01", "to": "2026-06-30", "position": "All"},
        )
        export_response = self._export_response(position="All")

        self.assertEqual(report_response.status_code, 200)
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(self.service.calls, 2)
        self.assertEqual(self.service.arguments[0], self.service.arguments[1])
        self.assertEqual(self.service.arguments[0]["position"], "All")

    def test_workbook_structure_and_columns(self):
        response = self._export_response()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        self.addCleanup(workbook.close)
        self.assertEqual(
            workbook.sheetnames,
            ["Cockpit Detail", "Cabin Detail", "Summary", "Report Information"],
        )
        self.assertEqual(
            tuple(cell.value for cell in workbook["Cockpit Detail"][4]),
            (
                "Position type",
                "Name",
                "Date",
                "Aircraft",
                "Flight number",
                "ADEP",
                "ADES",
                "OFF",
                "ON",
                "Block time",
                "Augmented (Heavy)",
            ),
        )
        self.assertEqual(
            tuple(cell.value for cell in workbook["Summary"][3]),
            ("Name", "Position type", "Flight count", "Official total"),
        )

    def test_detail_rows_are_grouped_by_crew_and_raw_position_is_preserved(self):
        response = self._export_response()
        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        self.addCleanup(workbook.close)
        worksheet = workbook["Cockpit Detail"]

        crew_headers = [
            worksheet.cell(row=row, column=1).value
            for row in range(1, worksheet.max_row + 1)
            if isinstance(worksheet.cell(row=row, column=1).value, str)
            and worksheet.cell(row=row, column=1).value.startswith("Crew member:")
        ]
        self.assertEqual(crew_headers, ["Crew member: =cmd|' /c calc'!A0"])
        self.assertEqual(worksheet["A8"].value, "PAD")
        self.assertEqual(worksheet["B8"].value, "'=cmd|' /c calc'!A0")

        cabin = workbook["Cabin Detail"]
        cabin_headers = [
            cabin.cell(row=row, column=1).value
            for row in range(1, cabin.max_row + 1)
            if isinstance(cabin.cell(row=row, column=1).value, str)
            and cabin.cell(row=row, column=1).value.startswith("Crew member:")
        ]
        self.assertEqual(cabin_headers, ["Crew member: Cabin Crew"])
        self.assertEqual(cabin["B8"].value, "Cabin Crew")

    def test_augmented_heavy_is_detail_only_and_renders_all_three_states(self):
        report = _report()
        report.crew_members[0].flights[0].augmented_heavy = True
        report.crew_members[0].flights[1].augmented_heavy = False
        report.crew_members[1].flights[0].augmented_heavy = None
        self.service.result = report

        response = self._export_response()
        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        self.addCleanup(workbook.close)

        cockpit = workbook["Cockpit Detail"]
        cabin = workbook["Cabin Detail"]
        summary = workbook["Summary"]
        self.assertEqual(cockpit["K4"].value, "Augmented (Heavy)")
        self.assertEqual(cockpit["K8"].value, "Yes")
        self.assertEqual(cockpit["K9"].value, "No")
        self.assertEqual(cabin["K8"].value, "Unknown")
        self.assertNotIn("Augmented (Heavy)", [cell.value for cell in summary[3]])

    def test_official_totals_and_missing_totals_are_not_recomputed(self):
        response = self._export_response()
        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        self.addCleanup(workbook.close)
        summary = workbook["Summary"]
        official_totals = {
            summary.cell(row=row, column=1).value: summary.cell(row=row, column=4).value
            for row in range(4, summary.max_row + 1)
            if summary.cell(row=row, column=1).value is not None
        }
        self.assertEqual(official_totals["'=cmd|' /c calc'!A0"], timedelta(hours=75, minutes=5))
        self.assertEqual(official_totals["Cabin Crew"], "Not available")
        self.assertEqual(official_totals["Maintenance Crew"], timedelta(hours=1))
        self.assertEqual(official_totals["Unclassified Crew"], "Not available")
        formulas = [
            cell.value
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        ]
        self.assertEqual(formulas, [])

    def test_official_total_duration_is_numeric_and_uses_elapsed_time_format(self):
        response = self._export_response()
        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        self.addCleanup(workbook.close)
        expected_days = (75 * 60 + 5) / (24 * 60)
        for cell in (workbook["Cockpit Detail"]["B7"], workbook["Summary"]["D4"]):
            self.assertEqual(cell.number_format, "[h]:mm")
            self.assertIsInstance(cell.value, (int, float, timedelta))
            self.assertNotIsInstance(cell.value, str)
            duration_days = (
                cell.value.total_seconds() / (24 * 60 * 60)
                if isinstance(cell.value, timedelta)
                else cell.value
            )
            self.assertAlmostEqual(duration_days, expected_days, places=8)

        missing_total = workbook["Cabin Detail"]["B7"]
        self.assertEqual(missing_total.value, "Not available")
        self.assertIsInstance(missing_total.value, str)
        self.assertEqual(missing_total.number_format, "General")

    def test_trn_and_psn_are_textual_or_unavailable_not_numeric_summary_totals(self):
        report = _report()
        report.crew_members.extend(
            [
                CrewMemberSummary(
                    crew_id="TRAINING-CREW",
                    person_code="TRAINING-CODE",
                    display_name="Training Crew",
                    full_name="Training Crew",
                    position_type="Cockpit",
                    status="TRN",
                    official_total="TRN",
                    raw_official_total="TRN",
                    flight_count=0,
                    flights=[],
                ),
                CrewMemberSummary(
                    crew_id="PASSENGER-CREW",
                    person_code="PASSENGER-CODE",
                    display_name="Passenger Crew",
                    full_name="Passenger Crew",
                    position_type="Cockpit",
                    official_total=None,
                    flight_count=1,
                    flights=[
                        _flight(
                            "PASSENGER-FLIGHT",
                            position="PSN",
                            block_time="02:00",
                            flight_number="RS-PSN",
                            journey_log={
                                "credentials": "CREDENTIAL-SENTINEL",
                                "jwt": "JWT-SENTINEL",
                                "raw_payload": "RAW-PAYLOAD-SENTINEL",
                            },
                        )
                    ],
                ),
            ]
        )
        self.service.result = report

        response = self._export_response()
        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        self.addCleanup(workbook.close)
        summary = workbook["Summary"]
        totals = {
            summary.cell(row=row, column=1).value: summary.cell(row=row, column=4)
            for row in range(4, summary.max_row + 1)
            if summary.cell(row=row, column=1).value is not None
        }

        self.assertEqual(totals["Training Crew"].value, "TRN")
        self.assertEqual(totals["Training Crew"].number_format, "General")
        self.assertEqual(totals["Passenger Crew"].value, "Not available")
        self.assertEqual(totals["Passenger Crew"].number_format, "General")
        duration_note = workbook["Report Information"]["B9"].value
        self.assertIn("not-a-duration", duration_note)
        self.assertNotIn("TRN", duration_note)

        workbook_text = "\n".join(
            str(cell.value)
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        for forbidden in (
            "journey_log",
            "credentials",
            "CREDENTIAL-SENTINEL",
            "JWT-SENTINEL",
            "RAW-PAYLOAD-SENTINEL",
        ):
            self.assertNotIn(forbidden, workbook_text)

    def test_formula_injection_is_escaped_and_unparsed_duration_is_reported(self):
        response = self._export_response()
        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        self.addCleanup(workbook.close)
        self.assertEqual(workbook["Summary"]["A4"].value, "'=cmd|' /c calc'!A0")
        information_values = [
            cell.value
            for row in workbook["Report Information"].iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertTrue(
            any(
                isinstance(value, str) and "not-a-duration" in value
                for value in information_values
            )
        )

    def test_filename_is_server_generated_and_safe(self):
        response = self._export_response(
            position="Cockpit;../secrets",
            crew_member="=cmd|' /c calc'!A0",
        )
        disposition = response.headers["content-disposition"]
        filename = re.search(r'filename="([^"]+)"', disposition).group(1)
        self.assertEqual(
            disposition,
            'attachment; filename="crew-hours-2026-06-01-to-2026-06-30.xlsx"',
        )
        self.assertRegex(filename, r"^[A-Za-z0-9._-]+$")
        self.assertNotIn("secrets", filename)
        self.assertNotIn("calc", filename)

    def test_content_disposition_is_exposed_to_allowed_browser_origin(self):
        response = self.client.get(
            "/api/statistics/crew-hours/report/export",
            params={"from": "2026-06-01", "to": "2026-06-30"},
            headers={"Origin": CORS_ORIGINS[0]},
        )

        self.assertEqual(response.status_code, 200)
        exposed_headers = {
            value.strip().lower()
            for value in response.headers["access-control-expose-headers"].split(",")
        }
        self.assertIn("content-disposition", exposed_headers)

    def test_report_information_has_period_timestamp_user_and_source(self):
        response = self._export_response()
        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        self.addCleanup(workbook.close)
        information = workbook["Report Information"]
        values = [
            cell.value
            for row in information.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertIn("2026-06-01", values)
        self.assertIn("2026-06-30", values)
        self.assertIn("exporter@example.com", values)
        self.assertIn("LEON MCP", values)
        generated_at = information["B5"].value
        self.assertIsInstance(generated_at, str)
        self.assertRegex(generated_at, r"^2026-.*Z$")

    def test_forbidden_raw_payload_and_internal_identifiers_are_absent(self):
        response = self._export_response()
        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        self.addCleanup(workbook.close)
        values = [
            str(cell.value)
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        workbook_text = "\n".join(values)
        for forbidden in (
            "RAW LEON PAYLOAD",
            "INTERNAL-FLIGHT-001",
            "INTERNAL-CREW-001",
            "INTERNAL-CODE-001",
            "journey_log",
            "flight_nid",
            "crew_id",
            "person_code",
        ):
            self.assertNotIn(forbidden, workbook_text)

    def test_leon_error_mapping_matches_report_endpoint(self):
        cases = (
            (LeonTimeoutError("timeout sentinel"), 504, "LEON report request timed out."),
            (LeonRateLimitError(17), 429, "LEON report rate limit exceeded."),
        )
        for error, expected_status, expected_detail in cases:
            with self.subTest(error=type(error).__name__):
                self.service = _RecordingReportService(error=error)
                app.dependency_overrides[get_crew_hours_service] = lambda: self.service
                response = self._export_response()
                self.assertEqual(response.status_code, expected_status)
                self.assertEqual(response.json(), {"detail": expected_detail})
                self.assertEqual(self.service.calls, 1)


if __name__ == "__main__":
    unittest.main()
